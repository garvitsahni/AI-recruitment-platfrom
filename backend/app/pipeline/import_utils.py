import csv
import datetime
import os
import re
import zipfile

import openpyxl


# ── File reading helpers ──────────────────────────────────────────────────

def _read_as_tsv(path: str) -> list[tuple]:
    """
    Read a file as tab-separated values.
    IHMCL portal exports data as TSV but names it .xls — this handles that.
    """
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f, delimiter="\t")
        return [tuple(row) for row in reader]


def _read_as_xlsx(path: str) -> list[tuple]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _read_as_xls_binary(path: str) -> list[tuple]:
    """Try real binary .xls via xlrd (genuine Excel 97-2003 files)."""
    import xlrd  # type: ignore
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    rows = []
    for rx in range(ws.nrows):
        row = []
        for cx in range(ws.ncols):
            cell = ws.cell(rx, cx)
            if cell.ctype == 2:
                val = cell.value
                row.append(int(val) if val == int(val) else val)
            elif cell.ctype == 0:
                row.append(None)
            else:
                row.append(str(cell.value))
        rows.append(tuple(row))
    return rows


def _load_rows(path: str) -> list[tuple]:
    """
    Auto-detect file format. IHMCL portal .xls exports are TSV in disguise —
    try TSV first, fall back to real binary xls, then xlsx.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        return _read_as_xlsx(path)

    if ext == ".xls":
        # Primary: TSV (IHMCL portal export format)
        try:
            rows = _read_as_tsv(path)
            if rows and len(rows[0]) > 3:
                return rows
        except Exception:
            pass
        # Fallback: real binary xls
        try:
            return _read_as_xls_binary(path)
        except Exception:
            pass
        # Last resort: try as xlsx
        return _read_as_xlsx(path)

    # .csv or unknown: treat as TSV/CSV
    return _read_as_tsv(path)


# ── Column / cell helpers ─────────────────────────────────────────────────

def _find_col(headers: list[str], *names: str) -> int | None:
    for name in names:
        nl = name.lower()
        for i, h in enumerate(headers):
            if nl in h:
                return i
    return None


def _cell(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


# ── Date parsing ──────────────────────────────────────────────────────────

def _parse_date(val: str) -> datetime.date | None:
    if not val or str(val).strip().lower() in ("", "n/a", "none", "null"):
        return None
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try just the first 10 chars (handles "2022-07-11T00:00:00" etc.)
    if len(s) > 10:
        return _parse_date(s[:10])
    return None


# ── Experience calculation ────────────────────────────────────────────────

def _calc_experience_years(row: tuple, headers: list[str]) -> float:
    """
    Sum up work experience from OrgName1-8 / FromDate1-8 / ToDate1-8.
    An empty ToDate means the candidate is currently employed there.
    """
    today = datetime.date.today()
    total_days = 0

    for i in range(1, 9):
        org_col  = _find_col(headers, f"org name{i}")
        from_col = _find_col(headers, f"from date{i}")
        to_col   = _find_col(headers, f"to date{i}")

        if org_col is None:
            continue
        org_val = _cell(row, org_col)
        if not org_val or org_val.lower() in ("n/a", "", "none"):
            continue

        from_date = _parse_date(_cell(row, from_col)) if from_col is not None else None
        to_val    = _cell(row, to_col) if to_col is not None else None
        to_date   = _parse_date(to_val) if to_val else None

        if from_date is None:
            continue
        if to_date is None or to_date > today:
            to_date = today

        delta = (to_date - from_date).days
        if delta > 0:
            total_days += delta

    return round(total_days / 365.25, 1)


# ── Education helpers ─────────────────────────────────────────────────────

def _highest_degree_and_education(row: tuple, headers: list[str]) -> tuple[str, str]:
    """
    Determine the highest qualification: PhD > PG > Graduation > 12th.
    Returns (degree_label, one_line_summary).
    """
    def get(col_name: str) -> str:
        idx = _find_col(headers, col_name)
        return _cell(row, idx) if idx is not None else ""

    def fmt(course: str, spec: str, univ: str) -> str:
        parts = [course.strip()]
        if spec and spec.strip().lower() not in ("n/a", ""):
            parts.append(spec.strip())
        if univ and univ.strip().lower() not in ("n/a", ""):
            parts.append(univ.strip())
        return ", ".join(parts)

    if get("phd institution name"):
        return "PhD", fmt(
            get("phd course name") or "PhD",
            get("phd specialization"),
            get("phd board university") or get("phd institution name"),
        )

    if get("p institution name"):
        course = get("p course name") or "Post-Graduation"
        return course.strip(), fmt(
            course, get("p specialization"),
            get("p board university") or get("p institution name"),
        )

    if get("g institution name"):
        course = get("g course name") or "Graduation"
        return course.strip(), fmt(
            course, get("g specialization"),
            get("g board university") or get("g institution name"),
        )

    if get("12 institution name"):
        board = get("12 board university") or get("12 institution name")
        return "Class XII", f"Class XII, {board}"

    return "", ""


def _extract_skills(row: tuple, headers: list[str]) -> list[str]:
    skills = []
    for i in range(1, 6):
        idx = _find_col(headers, f"tech name{i}")
        if idx is not None:
            val = _cell(row, idx)
            if val and val.lower() not in ("n/a", "", "none"):
                skills.append(val.strip())
    return skills


# ── Public API ────────────────────────────────────────────────────────────

def parse_candidate_excel(path: str) -> list[dict]:
    """
    Reads a recruiter-provided candidate sheet. Handles:
      - IHMCL portal TSV exports saved as .xls  (primary format)
      - Real .xlsx files
      - Real binary .xls files

    Returns a list of candidate dicts with:
      reference_number, candidate_name, candidate_email,
      cv_url  — direct URL to the candidate's CV PDF,
      parsed_form_data — pre-extracted structured data so Stage 2 (Gemini)
                         is skipped for these candidates.
    """
    rows = _load_rows(path)
    if not rows:
        return []

    raw_headers = rows[0]
    headers = [str(h).strip().lower() if h is not None else "" for h in raw_headers]

    ref_col     = _find_col(headers, "id.", "id ", "reference", "ref no", "ref_no", "application id", "sno", "s.no")
    name_col    = _find_col(headers, "name")
    email_col   = _find_col(headers, "email")
    dob_col     = _find_col(headers, "dob")
    aadhaar_col = _find_col(headers, "aadhaar")
    cv_col      = _find_col(headers, "user cv file", "cv file", "resume url", "resume file")

    candidates = []
    for idx, row in enumerate(rows[1:], start=1):
        if row is None or all((c is None or str(c).strip() == "") for c in row):
            continue

        ref    = _cell(row, ref_col)   or f"APP{idx:04d}"
        name   = _cell(row, name_col)  or ""
        email  = _cell(row, email_col) or ""
        cv_url = _cell(row, cv_col)    or None

        # Format DOB as DD-MM-YYYY for frontend display
        dob_raw = _cell(row, dob_col) or None
        dob = None
        if dob_raw:
            d = _parse_date(dob_raw)
            dob = d.strftime("%d-%m-%Y") if d else dob_raw

        highest_degree, education_summary = _highest_degree_and_education(row, headers)
        experience_years = _calc_experience_years(row, headers)
        skills = _extract_skills(row, headers)
        aadhaar = _cell(row, aadhaar_col) or None

        parsed_form_data = {
            "dob": dob,
            "education": education_summary or None,
            "highest_degree": highest_degree or None,
            "experience_years": experience_years,
            "skills": skills,
            "idNumber": aadhaar,
        }

        candidates.append({
            "reference_number": ref,
            "candidate_name": name,
            "candidate_email": email,
            "cv_url": cv_url,
            "parsed_form_data": parsed_form_data,
        })

    return candidates


def extract_zip(zip_path: str, dest_dir: str) -> list[str]:
    """Extracts a ZIP of candidate documents and returns the list of extracted file paths."""
    os.makedirs(dest_dir, exist_ok=True)
    extracted = []
    with zipfile.ZipFile(zip_path) as zf:
        for member in zf.namelist():
            if member.endswith("/"):
                continue
            target = os.path.join(dest_dir, os.path.basename(member))
            with zf.open(member) as src, open(target, "wb") as dst:
                dst.write(src.read())
            extracted.append(target)
    return extracted


def match_resume_for_candidate(files: list[str], reference_number: str, candidate_name: str) -> str | None:
    """Best-effort match between a candidate and one of the extracted PDF files."""
    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", (s or "").lower())

    ref_n       = norm(reference_number)
    name_n      = norm(candidate_name)
    name_tokens = [norm(t) for t in (candidate_name or "").split() if len(t) > 2]
    pdf_files   = [f for f in files if f.lower().endswith(".pdf")]

    def score(fname: str) -> int:
        base = norm(os.path.basename(fname))
        s = 0
        if ref_n  and ref_n  in base: s += 5
        if name_n and name_n in base: s += 5
        s += sum(1 for t in name_tokens if t in base)
        if "resume" in base or "cv" in base: s += 2
        return s

    if not pdf_files:
        return None
    best = max(pdf_files, key=score)
    return best if score(best) > 0 else pdf_files[0]
